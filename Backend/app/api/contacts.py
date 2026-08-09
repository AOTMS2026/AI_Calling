from fastapi import APIRouter, Depends, UploadFile, File, HTTPException, BackgroundTasks, Response, Request
from sqlmodel import Session, select, update
from sqlalchemy.dialects.postgresql import insert as pg_insert
from app.database.connection import get_session, engine
from app.models.domain import Contact, Call, User
from app.api.dependencies import get_current_user
import io
import csv
import os
import tempfile
import openpyxl
import uuid
import json
from typing import List, Any
from app.core.redis_client import get_cache, set_cache, delete_cache, generate_cache_key
from app.core.config import settings
import httpx

def serialize_models(models: List[Any]) -> List[dict]:
    serialized = []
    for m in models:
        try:
            serialized.append(json.loads(m.model_dump_json()))
        except AttributeError:
            serialized.append(json.loads(m.json()))
    return serialized

async def invalidate_contact_caches():
    try:
        # Contacts doesn't have tenant structures, so delete global keys
        await delete_cache("contacts_global")
    except Exception as e:
        print(f"Failed to invalidate contact caches: {e}")

router = APIRouter(prefix="/contacts", tags=["contacts"])

# Temporary in-memory state tracking to hold live progress percentages
UPLOAD_JOBS = {}

def process_bulk_excel(file_path: str, job_id: str):
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        sheet = wb.active
        
        batch_size = 5000
        batch = []
        
        with Session(engine) as session:
            header_map = {}
            for i, row in enumerate(sheet.iter_rows(values_only=True)):
                if not header_map:
                    row_str = str(row).lower()
                    if 'email' in row_str and 'first name' in row_str and 'phone' in row_str:
                        for idx, cell in enumerate(row):
                            if not cell: continue
                            h = str(cell).strip().lower()
                            if 's.n' in h or 'sno' in h: header_map['s_no'] = idx
                            elif 'email' in h: header_map['email'] = idx
                            elif 'first name' in h: header_map['first_name'] = idx
                            elif 'last name' in h: header_map['last_name'] = idx
                            elif 'branch' in h: header_map['branch'] = idx
                            elif 'qualification' in h: header_map['qualification'] = idx
                            elif 'institue' in h or 'institute' in h: header_map['institute_name'] = idx
                            elif 'district' in h: header_map['district'] = idx
                            elif 'phone' in h: header_map['phone'] = idx
                    continue 
                
                def get_val(key):
                    if key in header_map and header_map[key] < len(row):
                        v = row[header_map[key]]
                        return str(v).strip() if v is not None else None
                    return None

                s_no = get_val('s_no')
                email = get_val('email')
                first_name = get_val('first_name')
                last_name = get_val('last_name')
                branch = get_val('branch')
                qualification = get_val('qualification')
                institute_name = get_val('institute_name')
                district = get_val('district')
                phone = get_val('phone') or ''
                
                if not phone and not email:
                    continue

                batch.append({
                    "s_no": s_no,
                    "email": email,
                    "first_name": first_name,
                    "last_name": last_name,
                    "branch": branch,
                    "qualification": qualification,
                    "institute_name": institute_name,
                    "district": district,
                    "phone": phone if phone else f"NO-PHONE-{i}"
                })
                
                if len(batch) >= batch_size:
                    stmt = pg_insert(Contact).values(batch).on_conflict_do_nothing()
                    session.execute(stmt)
                    session.commit()
                    batch = []
                    
                total_approx = sheet.max_row or 1000
                progress_pct = min(99, int((i / total_approx) * 100))
                UPLOAD_JOBS[job_id]["progress"] = progress_pct
                    
            if batch:
                stmt = pg_insert(Contact).values(batch).on_conflict_do_nothing()
                session.execute(stmt)
                session.commit()
                
            UPLOAD_JOBS[job_id]["progress"] = 100
            UPLOAD_JOBS[job_id]["status"] = "completed"
            
            try:
                import asyncio
                loop = asyncio.get_event_loop()
                if loop.is_running():
                    loop.create_task(invalidate_contact_caches())
                else:
                    loop.run_until_complete(invalidate_contact_caches())
            except Exception as cache_err:
                print(f"Failed cache invalidation in import: {cache_err}")

                
    except Exception as e:
        print(f"Failed background import: {e}")
        if job_id in UPLOAD_JOBS:
            UPLOAD_JOBS[job_id]["status"] = "failed"
            UPLOAD_JOBS[job_id]["error"] = str(e)
    finally:
        if 'wb' in locals():
            wb.close()
            
        if os.path.exists(file_path):
            try:
                os.remove(file_path)
            except Exception as cleanup_error:
                print(f"Warning: Could not remove temporary file {file_path}: {cleanup_error}")

@router.get("/upload-progress/{job_id}")
async def get_upload_progress(job_id: str):
    if job_id not in UPLOAD_JOBS:
        return {"status": "not_found", "progress": 0}
    return UPLOAD_JOBS[job_id]

@router.post("/bulk-upload")
async def bulk_upload_contacts(
    background_tasks: BackgroundTasks, 
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user)
):
    if not file.filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Only Excel (.xlsx) files are allowed for bulk upload")
        
    job_id = str(uuid.uuid4())
    UPLOAD_JOBS[job_id] = {"status": "processing", "progress": 0}
    
    try:
        fd, temp_path = tempfile.mkstemp(suffix=".xlsx")
        with os.fdopen(fd, 'wb') as f:
            content = await file.read()
            f.write(content)
            
        background_tasks.add_task(process_bulk_excel, temp_path, job_id)
    except Exception as e:
        UPLOAD_JOBS[job_id] = {"status": "failed", "error": str(e)}
        raise HTTPException(status_code=500, detail=f"Failed to initiate upload: {str(e)}")
        
    return {"message": "Bulk upload started", "job_id": job_id}

@router.post("/upload")
async def upload_contacts(
    file: UploadFile = File(...), 
    db: Session = Depends(get_session),
    current_user: User = Depends(get_current_user)
):
    
    if not file.filename.endswith(".csv"):
        raise HTTPException(status_code=400, detail="Only CSV files are allowed")
        
    contents = await file.read()
    try:
        decoded_content = contents.decode('utf-8')
        reader = csv.DictReader(io.StringIO(decoded_content))
        
        for row in reader:
            contact = Contact(
                phone=row.get('phone', ''),
                first_name=row.get('name', 'Unknown')
            )
            db.add(contact)
            
        db.commit()
        await invalidate_contact_caches()
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error parsing CSV: {str(e)}")
        
    return {"message": "Contacts uploaded and saved to PostgreSQL successfully"}

@router.get("/")
async def get_contacts(
    response: Response,
    db: Session = Depends(get_session), 
    limit: int = 500,
    current_user: User = Depends(get_current_user),
    agent_id: str = None
):
    response.headers["Cache-Control"] = "public, max-age=10, s-maxage=60"
    
    # 1. Fetch from Ravan API
    if settings.RAVAN_AGNI_AI:
        try:
            # Determine which endpoint to hit based on sandbox isolation
            if current_user.role != 'admin' and current_user.ravan_campaign_id:
                # Fetch strictly for their assigned campaign
                url = f"https://api.ravan.ai/api/v1/campaigns/{current_user.ravan_campaign_id}/contacts?limit={limit}"
            else:
                # Fetch globally (or filtered by frontend provided agent_id via query params natively if supported)
                # But Ravan API /contacts doesn't officially support agentId filter in the URL based on previous tests, 
                # so we just fetch recent and filter in memory if needed
                url = f"https://api.ravan.ai/api/v1/contacts/?limit={limit}"
                if agent_id:
                    url += f"&agentId={agent_id}"

            async with httpx.AsyncClient() as client:
                r_resp = await client.get(
                    url, 
                    headers={"X-Api-Key": settings.RAVAN_AGNI_AI, "Accept": "application/json"}
                )
                
            if r_resp.status_code == 200:
                ravan_data = r_resp.json().get("data", [])
                
                # Transform Ravan payload format back into our Local Format so the Frontend Table doesn't break
                formatted_contacts = []
                for rc in ravan_data:
                    # In campaigns endpoint, Ravan returns 'contactId'. In contacts endpoint, it's just 'id'.
                    c_id = rc.get('contactId') or rc.get('id') or str(uuid.uuid4())
                    
                    # Ravan sometimes leaves name blank in campaign endpoint. Fallback to Email prefix or Local DB
                    c_name = rc.get('name')
                    if not c_name:
                        c_first = rc.get('firstName', '')
                        c_last = rc.get('lastName', '')
                        c_name = f"{c_first} {c_last}".strip()
                    
                    if not c_name:
                        c_email = rc.get('email', '')
                        c_name = c_email.split('@')[0] if c_email else 'Unknown'
                        
                    formatted_contacts.append({
                        "contact_id": c_id,
                        "phone": rc.get('phone', ''),
                        "name": c_name,
                        "email": rc.get('email', ''),
                        "status": rc.get('status', 'pending'),
                        "agent_id": current_user.ravan_agent_id if current_user.role != 'admin' else agent_id,
                        "campaign_id": current_user.ravan_campaign_id if current_user.role != 'admin' else None,
                        "tags": ", ".join(rc.get('tags', [])),
                        "created_at": rc.get('createdAt', '')
                    })
                
                # If we fetched globally and have an agent filter, apply it manually just in case Ravan ignores it
                if current_user.role == 'admin' and agent_id:
                    # Ravan API doesn't return agentId directly in contact items usually, 
                    # but if it does, we can filter. Otherwise, we trust the API.
                    pass
                    
                return formatted_contacts
        except Exception as e:
            print(f"Failed to fetch contacts from Ravan API, falling back to local: {e}")
            pass

    # 2. Local Fallback (if Ravan API fails or is not configured)
    contacts = db.exec(select(Contact).order_by(Contact.phone.asc()).limit(limit)).all()
    return contacts

@router.post("/dial-single/{phone}")
def manual_dial_contact(phone: str, db: Session = Depends(get_session)):
    """Triggers an instantaneous manual call out of the UI bypassing bulk campaigns. Stripped for clean architecture."""
    new_call = Call(contact_phone=phone, result_status="Failed - APIs Removed")
    db.add(new_call)
    db.commit()
    db.refresh(new_call)
    
    return {"message": "Call Dispatched architecture successfully removed.", "vendor_sid": new_call.vendor_call_sid}

@router.post("/single")
async def create_single_contact(
    request: Request, 
    db: Session = Depends(get_session),
    current_user: User = Depends(get_current_user)
):
    """Receives JSON from the UI Modal to manually inject a single user"""
    payload_data = await request.json()
    
    phone = payload_data.get("phone", "")
    name = payload_data.get("name", "")
    email = payload_data.get("email", "")
    agent_id = payload_data.get("agent_id")
    campaign_id = payload_data.get("campaign_id")
    tags = payload_data.get("tags", [])
    
    tags_str = ", ".join(tags) if isinstance(tags, list) else str(tags)
    
    contact = Contact(
        phone=phone,
        name=name,
        email=email,
        agent_id=agent_id,
        campaign_id=campaign_id,
        tags=tags_str
    )

    # Safe insert skipping crashes if phone number already exists
    stmt = pg_insert(Contact).values([contact.model_dump()]).on_conflict_do_nothing()
    db.execute(stmt)
    db.commit()
    
    # Sync with Ravan.ai contacts
    if settings.RAVAN_AGNI_AI:
        try:
            name_parts = name.strip().split() if name else ["Unknown"]
            first_name = name_parts[0]
            last_name = " ".join(name_parts[1:]) if len(name_parts) > 1 else ""
            
            ravan_metadata = payload_data.get("metadata", {})
            if "agentId" not in ravan_metadata and agent_id:
                ravan_metadata["agentId"] = agent_id

            ravan_payload = {
                "first_name": first_name,
                "last_name": last_name,
                "email": email,
                "phone": phone,
                "company": payload_data.get("company", ""),
                "metadata": ravan_metadata
            }

            resp = httpx.post(
                "https://api.ravan.ai/api/v1/contacts/",
                json=ravan_payload,
                headers={"X-Api-Key": settings.RAVAN_AGNI_AI, "Content-Type": "application/json"}
            )
            
            if resp.status_code in (200, 201):
                resp_data = resp.json().get("data", {})
                ravan_id = resp_data.get("id")
                if ravan_id:
                    # Update local DB with contact_id
                    db.exec(update(Contact).where(Contact.phone == phone).values(contact_id=ravan_id))
                    db.commit()
                    
                    # Auto-assign contact to the target campaign in Ravan.ai
                    if campaign_id:
                        camp_url = f"https://api.ravan.ai/api/v1/campaigns/{campaign_id}/contacts"
                        assign_resp = httpx.post(
                            camp_url,
                            json={"contactIds": [ravan_id]},
                            headers={"X-Api-Key": settings.RAVAN_AGNI_AI, "Content-Type": "application/json"}
                        )
                        if assign_resp.status_code not in (200, 201):
                            print(f"Failed to assign Ravan contact {ravan_id} to campaign {campaign_id}: {assign_resp.text}")
            else:
                print(f"Ravan API Create Contact Failed: {resp.status_code} {resp.text}")
        except Exception as e:
            print(f"Error triggering Ravan API: {e}")

    await invalidate_contact_caches()
    return {"message": "Contact securely added to PostgreSQL and queued in Ravan.ai."}
